#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""xlsx 写出库：把规范化数据按模板 schema 渲染成干净的 Excel（双行表头 + 动态日期区 + 样式）。

被 build_clean_templates.py（生成空白模板）与 fill_template.py（填充输出）共用。
设计要点：输出是“从 schema 现搭”的，天然不含示例数据，因此“删除模板示例数据”自动满足。
日期区按真实数据日期动态展开（按周 / 按日 / 按白夜班分组）。
"""
from __future__ import annotations
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 配色：克制的蓝灰，避免 AI-slop 紫渐变
C_TITLE   = "1F4E5F"   # 标题底
C_GROUP   = "2E6E7E"   # 分组表头底（周/交付日期）
C_LEAF    = "5B9AA8"   # 叶子表头底（日期/字段）
C_NOTE    = "FFF7E6"   # 备注行底
C_BORDER  = "BFC9CC"
THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
F_GROUP = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
F_LEAF  = Font(name="微软雅黑", size=9,  bold=True, color="FFFFFF")
F_NOTE  = Font(name="微软雅黑", size=9,  italic=True, color="8A6D3B")
F_DATA  = Font(name="微软雅黑", size=10, color="222222")
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _iso_week_label(d: dt.date, prefix="W") -> str:
    return f"{prefix}{d.isocalendar()[1]:02d}"


def build_date_leaves(dates: list[str], group_by: str, week_prefix="W",
                      header_format="%Y-%m-%d", shifts=None):
    """把排好序的 ISO 日期列表展开成叶子列。
    返回 [{top, bottom, key}] ；top=分组标签（同组会在表头合并），bottom=叶子标签。"""
    ds = sorted({d for d in dates if d})
    leaves = []
    for d in ds:
        try:
            day = dt.date.fromisoformat(d) if len(d) == 10 else None
        except ValueError:
            day = None
        if group_by == "iso_week" and day:
            top = _iso_week_label(day, week_prefix)
            bottom = f"{day.month}/{day.day}"
            leaves.append({"top": top, "bottom": bottom, "key": d})
        elif group_by == "day_shift" and day and shifts:
            for sh in shifts:
                leaves.append({"top": f"{day.month}/{day.day}", "bottom": sh,
                               "key": f"{d}|{sh}"})
        elif group_by == "day_shift" and day:
            leaves.append({"top": "交付日期", "bottom": f"{day.month}/{day.day}", "key": d})
        else:  # day
            label = f"{day.month}月{day.day}日" if day else d
            leaves.append({"top": "交付日期", "bottom": label, "key": d})
    return leaves


def build_leaf_columns(template: dict, dates: list[str]):
    """按模板 schema 的列顺序，生成最终叶子列清单。
    每个叶子: {canon, role, top, bottom, fill}  fill 决定取数来源。"""
    out = []
    dr = next((c for c in template["columns"] if c["role"] == "date_region"), None)
    for col in template["columns"]:
        role = col["role"]
        if role == "date_region":
            leaves = build_date_leaves(
                dates, col.get("group_by", "day"),
                col.get("week_label_prefix", "W"),
                shifts=col.get("shifts"))
            for lf in leaves:
                out.append({"canon": lf["key"], "role": "date", "top": lf["top"],
                            "bottom": lf["bottom"], "fill": ("series", lf["key"])})
        elif role == "summary" and col.get("formula") == "row_sum_dates":
            out.append({"canon": col["canonical"], "role": "summary",
                        "top": col["canonical"], "bottom": "",
                        "fill": ("formula", "row_sum_dates")})
        else:
            src = "scalars"
            if role == "identity":
                src = "identity"
            elif role == "remark":
                src = "remark"
            out.append({"canon": col["canonical"], "role": role,
                        "top": col["canonical"], "bottom": "",
                        "fill": (src, col["canonical"])})
    return out


def _val_for(leaf, row):
    src, key = leaf["fill"]
    if src == "series":
        return row.get("series", {}).get(key)
    if src == "identity":
        return row.get("identity", {}).get(key)
    if src == "remark":
        return row.get("remark", {}).get(key, row.get("scalars", {}).get(key))
    if src == "scalars":
        return row.get("scalars", {}).get(key)
    return None


def write_form(path: str, template: dict, rows: list[dict] | None,
               dates: list[str] | None = None, title_override: str | None = None,
               blank_placeholder: bool = False) -> dict:
    """渲染并保存一个 xlsx。
    rows: 规范化数据 [{identity:{}, scalars:{}, remark:{}, series:{iso:qty}}]；None=空白模板。
    dates: 日期区要展开的 ISO 日期列表；blank_placeholder=True 时给空白模板放 3 个占位日期列。
    返回 {path, sheet, n_rows, n_cols}。"""
    rows = rows or []
    if dates is None:
        ds = set()
        for r in rows:
            ds.update(k for k in r.get("series", {}).keys())
        dates = sorted(ds)
    if not dates and blank_placeholder:
        # 空白模板给一个示意日期区（标注“示意”，非真实数据）
        base = dt.date.today()
        dates = [(base + dt.timedelta(days=i)).isoformat() for i in range(3)]

    leaves = build_leaf_columns(template, dates)
    ncol = len(leaves)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template.get("sheet_name", "Sheet1")[:31]

    # row1 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, ncol))
    tc = ws.cell(row=1, column=1, value=title_override or template["title"])
    tc.font, tc.fill, tc.alignment = F_TITLE, PatternFill("solid", fgColor=C_TITLE), AL_C
    ws.row_dimensions[1].height = 26
    # row2 备注
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, ncol))
    note = template.get("title_note", "")
    if blank_placeholder:
        note += "  ｜ 日期数量区按导入数据的实际日期自动展开，此处为示意。"
    nc = ws.cell(row=2, column=1, value=note)
    nc.font, nc.fill, nc.alignment = F_NOTE, PatternFill("solid", fgColor=C_NOTE), AL_L
    ws.row_dimensions[2].height = 18

    HR_TOP, HR_BOT = 3, 4
    # 写叶子表头（先全部写 bottom，再处理 top 合并）
    for j, lf in enumerate(leaves, start=1):
        cl = get_column_letter(j)
        top_cell = ws.cell(row=HR_TOP, column=j, value=lf["top"])
        bot_cell = ws.cell(row=HR_BOT, column=j, value=lf["bottom"])
        for c in (top_cell, bot_cell):
            c.alignment, c.border = AL_C, BORDER
        if lf["bottom"] == "":
            # 非日期列：上下两行合并，蓝灰
            ws.merge_cells(start_row=HR_TOP, start_column=j, end_row=HR_BOT, end_column=j)
            top_cell.font = F_GROUP
            top_cell.fill = PatternFill("solid", fgColor=C_GROUP)
        else:
            top_cell.font, bot_cell.font = F_GROUP, F_LEAF
            top_cell.fill = PatternFill("solid", fgColor=C_GROUP)
            bot_cell.fill = PatternFill("solid", fgColor=C_LEAF)
        ws.column_dimensions[cl].width = _col_width(lf)

    # 合并连续同 top 的日期叶子（周/交付日期分组带）
    j = 1
    while j <= ncol:
        lf = leaves[j - 1]
        if lf["bottom"] != "" and lf["top"]:
            k = j
            while k + 1 <= ncol and leaves[k]["bottom"] != "" and leaves[k]["top"] == lf["top"]:
                k += 1
            if k > j:
                ws.merge_cells(start_row=HR_TOP, start_column=j, end_row=HR_TOP, end_column=k)
            j = k + 1
        else:
            j += 1
    ws.row_dimensions[HR_TOP].height = 18
    ws.row_dimensions[HR_BOT].height = 18

    # 数据行
    date_cols = [idx for idx, lf in enumerate(leaves, start=1) if lf["role"] == "date"]
    seq_col = next((idx for idx, lf in enumerate(leaves, start=1) if lf["canon"] == "序号"), None)
    r = HR_BOT + 1
    for n, row in enumerate(rows, start=1):
        for j, lf in enumerate(leaves, start=1):
            if lf["canon"] == "序号" and template_has_autoindex(template):
                val = n
            elif lf["fill"] == ("formula", "row_sum_dates"):
                if date_cols:
                    a = get_column_letter(date_cols[0]); b = get_column_letter(date_cols[-1])
                    val = f"=SUM({a}{r}:{b}{r})"
                else:
                    val = 0
            else:
                val = _val_for(lf, row)
            cell = ws.cell(row=r, column=j, value=val)
            cell.font, cell.border = F_DATA, BORDER
            cell.alignment = AL_L if lf["role"] in ("identity", "remark") else AL_C
        r += 1

    ws.freeze_panes = ws.cell(row=HR_BOT + 1, column=min(_freeze_col(leaves), ncol) + 1)
    wb.save(path)
    return {"path": path, "sheet": ws.title, "n_rows": len(rows), "n_cols": ncol,
            "leaves": leaves, "data_start_row": HR_BOT + 1,
            "header_top_row": HR_TOP, "header_bottom_row": HR_BOT}


def template_has_autoindex(template):
    return any(c.get("auto_index") for c in template["columns"])


def _col_width(lf):
    if lf["role"] == "date":
        return 8
    w = {"identity": 14, "remark": 18, "scalar": 10, "summary": 10}.get(lf["role"], 12)
    if lf["canon"] in ("规格型号", "物料描述", "描述", "物料名称"):
        return 26
    if lf["canon"] == "序号":
        return 6
    return w


def _freeze_col(leaves):
    # 冻结到第一个日期列之前
    for idx, lf in enumerate(leaves, start=1):
        if lf["role"] == "date":
            return idx - 1
    return min(4, len(leaves))
