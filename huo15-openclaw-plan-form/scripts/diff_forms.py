#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""比对【新文件】与【库里旧文件】，标注更新之处。文件名可不同——按模板主键(diff_key)匹配内容。

输入可为 .norm.json（extract 产物）或原始 .xls/.xlsx（脚本内部自动 extract）。

输出:
  1. 标注版 Excel：变化单元格红底 + 批注“原值: X”；新增行标识列绿底；新 sheet 列出“旧有新无”的删除行。
  2. 差异报告 JSON / 控制台摘要：新增 / 删除 / 变化单元格清单。

用法:
    python diff_forms.py --new <新文件|norm.json> --old <旧文件|norm.json>
        --template <模板id> [--out 标注版.xlsx] [--report diff.json]
"""
from __future__ import annotations
import sys, os, json, argparse, datetime as dt
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib_forms import load_schema, template_by_id
from lib_xlsx import write_form, build_leaf_columns
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

FILL_CHANGED = PatternFill("solid", fgColor="FFC7CE")   # 浅红：单元格有变化
FILL_ADDED   = PatternFill("solid", fgColor="C6EFCE")   # 浅绿：新增行
FILL_REMOVED = PatternFill("solid", fgColor="FFEB9C")   # 浅黄：删除行(单列表)


def load_as_norm(path, template):
    if path.lower().endswith(".norm.json") or path.lower().endswith(".json"):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    import extract_form
    return extract_form.extract(path, template)


def key_of(rec, keys):
    parts = [str(rec["identity"].get(k, "")).strip() for k in keys]
    return "|".join(parts) if any(parts) else None


def cell_values(rec):
    """把一条记录摊平成 {字段或日期: 值} 便于逐格比对。"""
    flat = {}
    for k, v in rec.get("identity", {}).items():
        flat[("id", k)] = v
    for k, v in rec.get("scalars", {}).items():
        flat[("sc", k)] = v
    for k, v in rec.get("series", {}).items():
        flat[("dt", k)] = v
    return flat


def norm_val(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def compute_diff(new_rows, old_rows, keys):
    new_map, old_map = {}, {}
    for r in new_rows:
        k = key_of(r, keys)
        if k:
            new_map.setdefault(k, r)
    for r in old_rows:
        k = key_of(r, keys)
        if k:
            old_map.setdefault(k, r)
    added = [k for k in new_map if k not in old_map]
    removed = [k for k in old_map if k not in new_map]
    changed = {}     # key -> [{field, old, new}]
    for k in new_map:
        if k not in old_map:
            continue
        nf, of = cell_values(new_map[k]), cell_values(old_map[k])
        diffs = []
        for fld in sorted(set(nf) | set(of), key=lambda x: (x[0], str(x[1]))):
            nv, ov = norm_val(nf.get(fld)), norm_val(of.get(fld))
            if nv != ov:
                diffs.append({"field": list(fld), "old": ov, "new": nv})
        if diffs:
            changed[k] = diffs
    return {"added": added, "removed": removed, "changed": changed,
            "new_map": new_map, "old_map": old_map}


def annotate(out_path, template, new_rows, diff, keys):
    dates = sorted({iso for r in new_rows for iso in r["series"].keys()})
    res = write_form(out_path, template, rows=new_rows, dates=dates)
    leaves = res["leaves"]
    data_r0 = res["data_start_row"]

    # 叶子 -> 列号
    col_of_field = {}    # ('id'/'sc', canon) -> col ; ('dt', iso) -> col
    for j, lf in enumerate(leaves, start=1):
        src, key = lf["fill"]
        if src == "series":
            col_of_field[("dt", key)] = j
        elif src in ("identity",):
            col_of_field[("id", lf["canon"])] = j
        elif src in ("scalars", "remark"):
            col_of_field[("sc", lf["canon"])] = j
        elif src == "formula":          # 汇总列(如 合计/已排数量)，源端值存在 scalars
            col_of_field[("sc", lf["canon"])] = j
    id_cols = [j for j, lf in enumerate(leaves, start=1) if lf["role"] == "identity"]

    wb = openpyxl.load_workbook(out_path)
    ws = wb[res["sheet"]]
    changed_cells = 0
    for i, r in enumerate(new_rows):
        row_no = data_r0 + i
        k = key_of(r, keys)
        if k in diff["added"]:
            for j in id_cols:
                ws.cell(row=row_no, column=j).fill = FILL_ADDED
        elif k in diff["changed"]:
            for d in diff["changed"][k]:
                fld = tuple(d["field"])
                col = col_of_field.get(fld)
                # remark 字段在布局里 src=scalars，已并入 ('sc',canon)
                if col is None and fld[0] == "sc":
                    col = col_of_field.get(("sc", fld[1]))
                if col:
                    c = ws.cell(row=row_no, column=col)
                    c.fill = FILL_CHANGED
                    c.comment = Comment(f"原值: {d['old'] or '（空）'}", "plan-form")
                    changed_cells += 1

    # 删除行（旧有新无）单列一个 sheet
    if diff["removed"]:
        ws2 = wb.create_sheet("本次删除(旧有新无)")
        ws2.append(["主键(" + "+".join(keys) + ")"] + keys + ["旧值摘要"])
        for k in diff["removed"]:
            rec = diff["old_map"][k]
            summary = "; ".join(f"{kk}={vv}" for kk, vv in list(rec["identity"].items())[:4])
            row = [k] + [rec["identity"].get(kk, "") for kk in keys] + [summary]
            ws2.append(row)
            for c in ws2[ws2.max_row]:
                c.fill = FILL_REMOVED

    # 顶部备注追加图例（写到标题备注行末尾的下一空列不可靠，改加批注到 A2）
    legend = (f"差异标注：红底=数值有变化(批注含原值, 共{changed_cells}格)；"
              f"绿底=新增行({len(diff['added'])})；黄色 sheet=删除行({len(diff['removed'])})。")
    a2 = ws.cell(row=2, column=1)
    a2.comment = Comment(legend, "plan-form")
    wb.save(out_path)
    return {"changed_cells": changed_cells, "path": out_path}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--new", required=True)
    ap.add_argument("--old", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", default=None)
    ap.add_argument("--report", default=None)
    args = ap.parse_args()

    schema = load_schema()
    tpl = template_by_id(schema, args.template)
    nnew = load_as_norm(args.new, tpl)
    nold = load_as_norm(args.old, tpl)
    new_rows = nnew.get("rows", [])
    old_rows = nold.get("rows", [])
    keys = tpl.get("diff_key", [])

    diff = compute_diff(new_rows, old_rows, keys)
    gen_ts = dt.datetime.now().strftime("%Y%m%d-%H%M")     # 生成时刻，精确到分钟
    out = args.out or f"差异标注_{tpl['display_name']}_{gen_ts}.xlsx"
    ann = annotate(out, tpl, new_rows, diff, keys)

    print(f"🔍 差异比对【{tpl['display_name']}】 主键={keys}")
    print(f"   新文件: {nnew.get('source_file', args.new)}  ({len(new_rows)} 行)")
    print(f"   旧文件: {nold.get('source_file', args.old)}  ({len(old_rows)} 行)")
    print(f"   ➕ 新增行: {len(diff['added'])}")
    print(f"   ➖ 删除行: {len(diff['removed'])}")
    print(f"   ✏️  变化行: {len(diff['changed'])}  (共 {ann['changed_cells']} 个单元格变化)")
    for k in list(diff["changed"])[:8]:
        ds = diff["changed"][k]
        shown = ", ".join(f"{d['field'][1]}:{d['old']}→{d['new']}" for d in ds[:4])
        print(f"      · {k}: {shown}{' …' if len(ds) > 4 else ''}")
    print(f"   ✓ 标注版已生成: {out}")

    report = {"template": tpl["id"], "keys": keys,
              "added": diff["added"], "removed": diff["removed"],
              "changed": diff["changed"], "changed_cells": ann["changed_cells"]}
    rep = args.report or (os.path.splitext(out)[0] + ".diff.json")
    with open(rep, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=1, default=str)
    print(f"   ✓ 差异报告: {rep}")


if __name__ == "__main__":
    main()
